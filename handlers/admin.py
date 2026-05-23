import csv
import io
import logging
from datetime import datetime

from aiogram import Router, types, F
from aiogram.filters import Command

from config import ADMIN_IDS
from database import get_all_users, get_users_count
from handlers.keyboards import admin_keyboard

router = Router()
logger = logging.getLogger(__name__)


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


@router.message(Command("admin"))
async def cmd_admin(message: types.Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emessiź!")
        return

    count = await get_users_count()
    await message.answer(
        f"🔐 <b>Admin Panel</b>\n\n"
        f"👥 Jámi registraciya: <b>{count}</b>\n"
        f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
        f"Tómendegi túymelerdi ishlatiń 👇",
        reply_markup=admin_keyboard(),
        parse_mode="HTML",
    )


@router.callback_query(F.data == "admin_refresh")
async def cb_refresh(callback: types.CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yoq!", show_alert=True)
        return
    count = await get_users_count()
    try:
        await callback.message.edit_text(
            f"🔐 <b>Admin Panel</b>\n\n"
            f"👥 Jámi registraciya: <b>{count}</b>\n"
            f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
            f"Tómendegi túymelerdi ishlatiń 👇",
            reply_markup=admin_keyboard(),
            parse_mode="HTML",
        )
    except Exception:
        pass
    await callback.answer("✅ Yangilandi!")


@router.callback_query(F.data == "admin_list")
async def cb_list_users(callback: types.CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yoq!", show_alert=True)
        return
    await callback.answer()
    users = await get_all_users()

    if not users:
        await callback.message.answer("📭 Hesh kim registraciyadan ótpegen.")
        return

    PAGE = 20
    chunk = users[:PAGE]
    lines = [f"📋 <b>Registraciyalar ro'yxati</b> (jámi: {len(users)})\n"]

    for i, u in enumerate(chunk, 1):
        # id, telegram_id, ism, familiya, tuman, maktab, til, telefon, sana
        sana = str(u[8])[:16] if u[8] else "—"
        lines.append(
            f"{i}. <b>{u[2]} {u[3]}</b>\n"
            f"   🏙️ {u[4]}\n"
            f"   🏫 {u[5]}\n"
            f"   📞 {u[7]} | 🌐 {u[6]}\n"
            f"   📅 {sana}\n"
        )

    if len(users) > PAGE:
        lines.append(f"\n<i>... yana {len(users) - PAGE} ta — CSV/Excel yuklab aliń</i>")

    await callback.message.answer("\n".join(lines), parse_mode="HTML")


@router.callback_query(F.data == "admin_csv")
async def cb_export_csv(callback: types.CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yoq!", show_alert=True)
        return
    await callback.answer("⏳ CSV tayyarlanmoqda...")
    users = await get_all_users()

    if not users:
        await callback.message.answer("📭 Hesh kim registraciyadan ótpegen.")
        return

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["#", "Telegram ID", "Atı", "Familiyası", "Tuman", "Mektep", "Til", "Telefon", "Sana"])

    for i, u in enumerate(users, 1):
        writer.writerow([i, u[1], u[2], u[3], u[4], u[5], u[6], u[7], u[8]])

    csv_bytes = buf.getvalue().encode("utf-8-sig")
    filename = f"registraciyalar_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    file = types.BufferedInputFile(csv_bytes, filename=filename)

    await callback.message.answer_document(
        file,
        caption=(
            f"📄 <b>CSV Export</b>\n"
            f"👥 Jámi: <b>{len(users)}</b> ta\n"
            f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ),
        parse_mode="HTML",
    )


@router.callback_query(F.data == "admin_excel")
async def cb_export_excel(callback: types.CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yoq!", show_alert=True)
        return
    await callback.answer("⏳ Excel tayyarlanmoqda...")
    users = await get_all_users()

    if not users:
        await callback.message.answer("📭 Hesh kim registraciyadan ótpegen.")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Registraciyalar"

        headers = ["#", "Telegram ID", "Atı", "Familiyası", "Tuman", "Mektep", "Til", "Telefon", "Sana"]
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 20

        for i, u in enumerate(users, 2):
            row_data = [i - 1, u[1], u[2], u[3], u[4], u[5], u[6], u[7], str(u[8])[:16] if u[8] else ""]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[i].height = 16

        col_widths = [5, 14, 15, 18, 20, 25, 20, 16, 20]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"registraciyalar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        file = types.BufferedInputFile(buf.read(), filename=filename)

        await callback.message.answer_document(
            file,
            caption=(
                f"📊 <b>Excel Export</b>\n"
                f"👥 Jámi: <b>{len(users)}</b> ta\n"
                f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ),
            parse_mode="HTML",
        )
    except ImportError:
        await callback.message.answer("❌ openpyxl o'rnatilmagan.")


@router.message(Command("stats"))
async def cmd_stats(message: types.Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer("❌ Siz admin emessiź!")
        return

    users = await get_all_users()
    count = len(users)

    lang_count: dict[str, int] = {}
    tuman_count: dict[str, int] = {}

    for u in users:
        # til
        lang = u[6] or "Belgisiz"
        lang_count[lang] = lang_count.get(lang, 0) + 1
        # tuman
        tuman = u[4] or "Belgisiz"
        tuman_count[tuman] = tuman_count.get(tuman, 0) + 1

    lang_lines = "\n".join(f"  • {k}: <b>{v}</b>" for k, v in lang_count.items())
    tuman_lines = "\n".join(f"  • {k}: <b>{v}</b>" for k, v in tuman_count.items())

    await message.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"👥 Jámi registraciya: <b>{count}</b>\n\n"
        f"🌐 Tillar boyınsha:\n{lang_lines or '  —'}\n\n"
        f"🏙️ Tumanlar boyınsha:\n{tuman_lines or '  —'}",
        parse_mode="HTML",
    )
